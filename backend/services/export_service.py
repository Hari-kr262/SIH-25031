"""Export service — PDF and Excel report generation."""

import io
from typing import List, Dict, Any, Optional
from datetime import datetime


class ExportService:
    """Generates PDF and Excel reports using ReportLab and openpyxl."""

    def generate_pdf_report(self, data: Dict[str, Any], title: str = "CivicResolve Report") -> bytes:
        """Generate a PDF report with issue statistics."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib import colors

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            # Title
            elements.append(Paragraph(title, styles["Title"]))
            elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
            elements.append(Spacer(1, 20))

            # Stats table
            if "stats" in data:
                table_data = [["Metric", "Value"]]
                for key, val in data["stats"].items():
                    table_data.append([str(key).replace("_", " ").title(), str(val)])

                table = Table(table_data, colWidths=[250, 200])
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
                ]))
                elements.append(table)

            doc.build(elements)
            return buffer.getvalue()
        except ImportError:
            return b"PDF generation requires reportlab package"

    def generate_excel_report(self, issues: List[Dict], sheet_name: str = "Issues") -> bytes:
        """Generate an Excel report from issue data."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            if not issues:
                wb_buffer = io.BytesIO()
                wb.save(wb_buffer)
                return wb_buffer.getvalue()

            # Headers
            headers = list(issues[0].keys())
            header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(col_idx)].width = 18

            # Data rows
            for row_idx, issue in enumerate(issues, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=str(issue.get(header, "")))

            wb_buffer = io.BytesIO()
            wb.save(wb_buffer)
            return wb_buffer.getvalue()
        except ImportError:
            return b"Excel generation requires openpyxl package"


export_service = ExportService()
